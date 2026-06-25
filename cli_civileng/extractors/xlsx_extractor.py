"""Extract structured project data from ArchiCAD SAF XLSX exports."""
import openpyxl

GROUND_LEVEL_THRESHOLD_M = 1.5  # Z height below which surfaces are considered ground level


def _polygon_area(coords: list[tuple[float, float, float]]) -> float:
    """Calculate 2D polygon area using shoelace formula (XY plane)."""
    if len(coords) < 3:
        return 0.0
    area = 0.0
    for i in range(len(coords)):
        x1, y1, _ = coords[i]
        x2, y2, _ = coords[(i + 1) % len(coords)]
        area += x1 * y2 - x2 * y1
    return abs(area) / 2


def _parse_nodes(workbook: openpyxl.Workbook) -> dict:
    """Parse StructuralPointConnection sheet into a {name: (x, y, z)} dict."""
    nodes: dict = {}
    if "StructuralPointConnection" not in workbook.sheetnames:
        return nodes

    ws = workbook["StructuralPointConnection"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] is not None:
            nodes[row[0]] = (
                float(row[1]),
                float(row[2]),
                float(row[3]) if row[3] else 0,
            )
    return nodes


def _parse_surfaces(
    workbook: openpyxl.Workbook, nodes: dict
) -> tuple[float, float, list[str], dict[str, int]]:
    """Parse StructuralSurfaceMember sheet.

    Returns (floor_areas_total, building_footprint_area, layers, elements_count).
    """
    floor_areas_total = 0.0
    footprint_areas: list[float] = []
    layers: set[str] = set()
    elements = {"slabs": 0}

    if "StructuralSurfaceMember" not in workbook.sheetnames:
        return 0.0, 0.0, [], elements

    ws = workbook["StructuralSurfaceMember"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        node_str = row[6] or ""
        layer = str(row[10]).strip() if row[10] else ""

        layers.add(layer)

        node_names = [n.strip() for n in str(node_str).split(";") if n.strip()]
        coords = [nodes[n] for n in node_names if n in nodes]

        if coords and len(coords) >= 3:
            area = _polygon_area(coords)

            # Classify: floors/slabs vs other
            if "Piso" in layer or "Laje" in layer:
                elements["slabs"] += 1
                floor_areas_total += area

                # Ground floor approximation: lowest Z surfaces
                z_avg = sum(c[2] for c in coords) / len(coords)
                if z_avg < GROUND_LEVEL_THRESHOLD_M:
                    footprint_areas.append(area)

    building_footprint_area = (
        round(sum(footprint_areas), 1) if footprint_areas else 0.0
    )
    return floor_areas_total, building_footprint_area, sorted(layers), elements


def _parse_curve_members(
    workbook: openpyxl.Workbook, layers: set[str], elements: dict[str, int]
) -> None:
    """Parse StructuralCurveMember sheet, updating layers and elements in place."""
    if "StructuralCurveMember" not in workbook.sheetnames:
        return

    ws = workbook["StructuralCurveMember"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and row[10]:
            element_type = str(row[1])
            layers.add(str(row[10]))
            if element_type == "Beam":
                elements["beams"] = elements.get("beams", 0) + 1
            elif element_type == "Column":
                elements["columns"] = elements.get("columns", 0) + 1


def extract_project_data(xlsx_path: str) -> dict:
    """Extract compliance-relevant data from ArchiCAD SAF XLSX.

    Returns dict with:
      - max_height: max Z coordinate from structural nodes
      - building_footprint_area: projection area from ground-floor surfaces
      - floor_areas_total: summed area of all floor/slab surfaces
      - layers: list of layer names found
      - structural_elements: counts of beams, columns, slabs
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Parse nodes
    nodes = _parse_nodes(wb)
    max_height = max(n[2] for n in nodes.values()) if nodes else 0.0

    # Parse surfaces
    floor_areas_total, building_footprint_area, surface_layers, elements = (
        _parse_surfaces(wb, nodes)
    )

    # Parse curve members
    layers_set = set(surface_layers)
    _parse_curve_members(wb, layers_set, elements)

    return {
        "building_footprint_area": round(building_footprint_area, 1),
        "floor_areas_total": round(floor_areas_total, 1),
        "max_height": round(max_height, 2),
        "layers": sorted(layers_set),
        "structural_elements": {
            "beams": elements.get("beams", 0),
            "columns": elements.get("columns", 0),
            "slabs": elements.get("slabs", 0),
        },
    }
